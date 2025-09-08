# useranme = "Isha"
# password = "12345"
#CRUD
import openpyxl
fp = r"D:\Isha\SeleniumPythonClass\External_files\TestData.xlsx"
wb = openpyxl.load_workbook(fp)
sheet_obj = wb["Data"]
rows = sheet_obj.max_row
columns = sheet_obj.max_column
print(rows)
print(columns)
# sheet_obj2 = wb["Sheet3"]
# print(sheet_obj2.max_row,sheet_obj2.max_column)

#r=3,c=2 (Suresh)   #retrive/get/read
user_name = sheet_obj.cell(9,2).value
print(user_name)

sheet_obj.cell(5,3).value = "c468"


sheet_obj.cell(8,1).value = "TC_7"

sheet_obj.cell(6,4).value = ""
wb.save(fp)
wb.close()





